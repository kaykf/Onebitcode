from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# 1-ler pasta de trabalho / planilha

wb = load_workbook("python_start/data2/pivot_table.xlsx")
sheet = wb["Relátorio"]

# 2- Referências de linhas e colunas 

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(
    sheet,
    min_col= min_column+1,
    max_col= max_column,
    min_row= min_row,
    max_row= max_row
)

categories = Reference(
    sheet,
    min_col= min_column,
    max_col=min_column,
    min_row=min_row+1,
    max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# 4- Criando o grafico

sheet.add_chart(barchart, "B10")
barchart.title= "Venda por Fabricantes"
barchart.style = 2

# 5- salvando workbook
wb.save('python_start/data2/barchart.xlsx')